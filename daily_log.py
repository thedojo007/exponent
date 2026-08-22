"""
daily_log.py — Pull tasks from ClickUp Outstanding List, generate printable Excel task log.

Setup:
    pip install requests openpyxl python-dotenv

Usage:
    python daily_log.py
    → Opens log_2026-08-21.xlsx in Excel. Edit, then print.

Requires CLICKUP_API_KEY in .env. ANTHROPIC_API_KEY is optional --
if unset, the leverage-picks section is skipped with a note in the
sheet rather than failing the run.
"""

import os
import json
import requests
from datetime import datetime, date
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

load_dotenv()

# --- Config ---
LIST_ID = "901715544578"
API_BASE = "https://api.clickup.com/api/v2"
STATE_FILE = Path(__file__).parent / "previous_pull.json"
OUTPUT_DIR = Path(r"D:\5. 미국 admin 일\4. 진행중\주간 업무 보고서")

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")

# --- Styles ---
THIN = Side(style="thin")
THICK = Side(style="medium")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_THICK = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
FONT_HEADER = Font(name="Malgun Gothic", size=12, bold=True)
FONT_SECTION = Font(name="Malgun Gothic", size=11, bold=True)
FONT_BODY = Font(name="Malgun Gothic", size=10)
FONT_TH = Font(name="Malgun Gothic", size=10, bold=True)
FONT_DAY_NORMAL = Font(name="Malgun Gothic", size=10)
FONT_DAY_TODAY = Font(name="Malgun Gothic", size=10, bold=True)
FONT_LEVERAGE = Font(name="Malgun Gothic", size=10)
FONT_LEVERAGE_HEADER = Font(name="Malgun Gothic", size=11, bold=True, color="1F5C2E")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

COL_WIDTHS = [38, 14, 12, 12, 10]
ROW_HEIGHT = 36  # Two-row equivalent height


# --- ClickUp API ---
def fetch_tasks():
    token = os.environ.get("CLICKUP_API_KEY")
    if not token:
        raise EnvironmentError("Set CLICKUP_API_KEY in your .env or environment.")

    headers = {"Authorization": token}
    tasks = []
    page = 0

    while True:
        resp = requests.get(
            f"{API_BASE}/list/{LIST_ID}/task",
            headers=headers,
            params={
                "statuses[]": ["backlog", "pushable", "on hold"],
                "subtasks": "false",
                "include_closed": "false",
                "page": str(page),
            },
        )
        resp.raise_for_status()
        data = resp.json()

        for t in data.get("tasks", []):
            pri = None
            if t.get("priority") and t["priority"].get("priority"):
                pri = t["priority"]["priority"]

            due = None
            if t.get("due_date"):
                due = datetime.fromtimestamp(int(t["due_date"]) / 1000).strftime("%m/%d")

            tasks.append({
                "name": t["name"],
                "status": t["status"]["status"].lower(),
                "priority": pri,
                "due_date": due,
            })

        if data.get("last_page", True):
            break
        page += 1

    return tasks


# --- State ---
def load_previous():
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    return None

def save_state(tasks):
    STATE_FILE.write_text(
        json.dumps(
            [{"name": t["name"], "status": t["status"]} for t in tasks],
            ensure_ascii=False, indent=2
        ),
        encoding="utf-8",
    )


# --- Helpers ---
PRIORITY_ORDER = {"urgent": 0, "high": 1, "normal": 2, "low": 3, None: 4}
PRIORITY_DISPLAY = {"urgent": "1", "high": "2", "normal": "3", "low": "4"}

def sort_key(t):
    return (PRIORITY_ORDER.get(t["priority"], 4), t["due_date"] or "99/99", t["name"].lower())

def pushable_yn(task, previous):
    if previous is None:
        return ""
    prev_map = {t["name"]: t["status"] for t in previous}
    if prev_map.get(task["name"]) == "pushable":
        return "Y"
    return ""


def set_col_widths(ws):
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_TH
        cell.border = BORDER_ALL
        cell.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = ROW_HEIGHT


def write_task_row(ws, row, task, previous, use_pushable=False):
    values = [
        task["name"],
        pushable_yn(task, previous) if use_pushable else "",
        task["due_date"] or "",
        "",
        PRIORITY_DISPLAY.get(task["priority"], ""),
    ]
    aligns = [ALIGN_LEFT, ALIGN_CENTER, ALIGN_CENTER, ALIGN_CENTER, ALIGN_CENTER]

    for col, (val, align) in enumerate(zip(values, aligns), 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = FONT_BODY
        cell.border = BORDER_ALL
        cell.alignment = align
    ws.row_dimensions[row].height = ROW_HEIGHT


def write_blank_row(ws, row):
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col, value="")
        cell.border = BORDER_ALL
    ws.row_dimensions[row].height = ROW_HEIGHT


# --- Leverage picks (isolated module -- own fetch input, own failure path,
#     never blocks the rest of the sheet from generating) ---

def call_claude(prompt: str, max_tokens: int = 500) -> str:
    url = "https://api.anthropic.com/v1/messages"
    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    body = {
        "model": "claude-sonnet-5",
        "max_tokens": max_tokens,
        "thinking": {"type": "disabled"},
        "messages": [{"role": "user", "content": prompt}],
    }
    resp = requests.post(url, headers=headers, json=body, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    return "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text")


def get_leverage_picks(tasks: list[dict]) -> str:
    """Pick 2 open items where Claude can carry the most of the actual
    work (drafting, templating) and Sungyun's remaining input is
    smallest. Hard-excludes financial authorization / judgment-call
    items in the prompt itself, not left to inference. Returns a
    display-ready string; never raises -- failures are caught and
    turned into a skip note so the rest of the sheet still builds."""
    if not ANTHROPIC_API_KEY:
        print("[DEBUG] ANTHROPIC_API_KEY missing -- skipping leverage picks")
        return "(Skipped: ANTHROPIC_API_KEY not set in .env.)"

    if not tasks:
        return "(No open items to evaluate.)"

    task_summary = "\n".join(f"- {t['name']}" for t in tasks)

    prompt = f"""Here are my open Outstanding List items (raw titles, often terse Korean shorthand):

{task_summary}

Pick exactly 2 where I (Claude) can do most of the actual work -- e.g.
drafting a reply/message, filling a template, writing a standard
document -- and Sungyun's remaining input is small (read, approve, send).

Hard exclude: anything requiring financial authorization, a wire/payment
decision, or judgment only he can make. If fewer than 2 genuinely qualify,
say so honestly instead of forcing a pick.

Output as plain short lines, no markdown bold/headers (this goes into an
Excel cell):
1. [task name] -- [what Claude would draft/do] -- [what's left for Sungyun]
2. [task name] -- [what Claude would draft/do] -- [what's left for Sungyun]
"""

    try:
        result = call_claude(prompt, max_tokens=500)
        print(f"[DEBUG] leverage picks returned: {repr(result)[:200]}")
        return result.strip()
    except requests.exceptions.RequestException as e:
        print(f"[DEBUG] leverage picks call failed: {e}")
        return f"(Skipped: Anthropic API call failed -- {e})"


def write_leverage_section(ws, row, picks_text):
    """Highlighted box near the top of Page 1 -- first thing seen each
    morning, separate from the task table below it."""
    ws.cell(row=row, column=1, value="Today's 2 Highest-Leverage Picks (low input / high Claude output):").font = FONT_LEVERAGE_HEADER
    row += 1

    cell = ws.cell(row=row, column=1, value=picks_text)
    cell.font = FONT_LEVERAGE
    cell.alignment = ALIGN_LEFT_TOP
    cell.border = BORDER_ALL
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=5)
    ws.row_dimensions[row].height = 20  # merged range absorbs the rest visually
    row += 3

    return row + 1  # trailing blank line before the next section


# --- Excel Generation ---
def generate_xlsx(tasks, previous):
    today = date.today()
    date_str = f"{today.month}/{today.day}/{today.year}"
    weekday = today.weekday()

    # Split tasks: backlog → not completed, pushable → to complete, on hold → things to keep in mind
    not_completed = sorted([t for t in tasks if t["status"] == "backlog"], key=sort_key)
    to_complete = sorted([t for t in tasks if t["status"] == "pushable"], key=sort_key)
    on_hold = [t for t in tasks if t["status"] == "on hold"]

    print("Getting leverage picks from Claude...")
    leverage_picks_text = get_leverage_picks(tasks)

    wb = Workbook()
    ws = wb.active
    ws.title = today.strftime("%#m.%d") if os.name == "nt" else today.strftime("%-m.%d")
    set_col_widths(ws)

    row = 1

    # === PAGE 1: Tasks not completed ===

    # Date
    ws.cell(row=row, column=1, value="Date:").font = FONT_HEADER
    row += 1
    ws.cell(row=row, column=1, value=date_str).font = FONT_BODY
    row += 2

    # Day of week
    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    for col, d in enumerate(days, 1):
        cell = ws.cell(row=row, column=col, value=d)
        cell.border = BORDER_ALL
        cell.alignment = ALIGN_CENTER
        if col - 1 == weekday and weekday < 5:
            cell.font = FONT_DAY_TODAY
            cell.border = BORDER_THICK
        else:
            cell.font = FONT_DAY_NORMAL
    row += 2

    # Today's 2 Highest-Leverage Picks -- placed first, before the task
    # table, so it's the first actionable thing seen on the printed page
    row = write_leverage_section(ws, row, leverage_picks_text)

    # Tasks not completed
    ws.cell(row=row, column=1, value="Tasks not completed:").font = FONT_SECTION
    row += 2
    headers = ["Task", "Pushable Y/N:", "Due Date?", "Completed?", "Priority"]
    write_header_row(ws, row, headers)
    row += 1

    for t in not_completed:
        write_task_row(ws, row, t, previous, use_pushable=True)
        row += 1

    for _ in range(3):
        write_blank_row(ws, row)
        row += 1

    # --- PAGE BREAK ---
    ws.row_breaks.append(Break(id=row - 1))

    # === PAGE 2: Tasks to complete ===
    row += 1
    ws.cell(row=row, column=1, value="Tasks to complete:").font = FONT_SECTION
    row += 2
    write_header_row(ws, row, headers)
    row += 1

    for t in to_complete:
        write_task_row(ws, row, t, previous, use_pushable=False)
        row += 1

    for _ in range(5):
        write_blank_row(ws, row)
        row += 1

    row += 1

    # --- Things to keep in mind (on hold items go here) ---
    ws.cell(row=row, column=1, value="Things to keep in mind:").font = FONT_SECTION
    row += 1

    for t in on_hold:
        cell = ws.cell(row=row, column=1, value=t["name"])
        cell.font = FONT_BODY
        cell.border = Border(bottom=THIN)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = ROW_HEIGHT
        row += 1

    # Extra blank lines for hand-writing
    for _ in range(4):
        ws.cell(row=row, column=1, value="").border = Border(bottom=THIN)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = ROW_HEIGHT
        row += 1

    row += 1

    # --- Things to note ---
    ws.cell(row=row, column=1, value="Things to note:").font = FONT_SECTION
    row += 1
    ws.cell(row=row, column=1, value="Order of Priority").font = Font(name="Malgun Gothic", size=10, bold=True)
    row += 2

    categories = ["지급 현황표", "영업부 자료", "장비 대여표", "주간업무 건"]
    for cat in categories:
        ws.cell(row=row, column=1, value=cat).font = Font(name="Malgun Gothic", size=10, bold=True)
        row += 1
        for _ in range(3):
            ws.cell(row=row, column=1, value="").border = Border(bottom=THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.row_dimensions[row].height = 24
            row += 1
        row += 1

    # --- Print settings ---
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return wb


# --- Main ---
def main():
    print("Pulling from ClickUp...")
    tasks = fetch_tasks()
    print(f"  {len(tasks)} tasks fetched")

    previous = load_previous()
    print(f"  Previous state: {'loaded' if previous else 'none (first run)'}")

    wb = generate_xlsx(tasks, previous)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"log_{date.today().isoformat()}.xlsx"
    out = OUTPUT_DIR / filename
    wb.save(str(out))

    save_state(tasks)
    print(f"  Saved → {out.resolve()}")

    import webbrowser
    webbrowser.open(str(out.resolve()))


if __name__ == "__main__":
    main()